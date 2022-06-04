from openpyxl import Workbook
from openpyxl.styles import Font
from Engine.CavingProductionPlanExtractionSpeedItem import CavingProductionPlanExtractionSpeedItem
from Engine.CavingProductionPlanTargetItem import CavingProductionPlanTargetItem
from Models.excel_utils import remove_default_worksheet
import numpy as np

CAVING_PLAN_CONFIGURATION_METADATA_SHEET = "Metadata"

CAVING_PLAN_CONFIGURATION_NAME_CELL = (1, 2)
CAVING_PLAN_CONFIGURATION_PERIOD_CELL = (2, 2)
CAVING_PLAN_CONFIGURATION_NUMBER_OF_SPEEDS_CELL = (3, 2)
CAVING_PLAN_CONFIGURATION_DENSITY_DATA_SET_CELL = (4, 2)
CAVING_PLAN_CONFIGURATION_AVERAGE_DATA_SET_CELL = (5, 2)
CAVING_PLAN_CONFIGURATION_SUMMATION_DATA_SET_CELL = (6, 2)


CAVING_PLAN_CONFIGURATION_TARGET_SHEET = "Target"

CAVING_PLAN_CONFIGURATION_DURATION_COLUMN = 4
CAVING_PLAN_CONFIGURATION_INCORPORATION_COLUMN = 3
CAVING_PLAN_CONFIGURATION_PERIOD_COLUMN = 1
CAVING_PLAN_CONFIGURATION_TARGET_COLUMN = 2

CAVING_PLAN_CONFIGURATION_SPEED_SHEET = "Speed"

CAVING_PLAN_CONFIGURATION_MINIMUM_PERCENTAGE_COLUMN = 1
CAVING_PLAN_CONFIGURATION_MAXIMUM_PERCENTAGE_COLUMN = 2
CAVING_PLAN_CONFIGURATION_SPEED_COLUMN = 3

CAVING_PLAN_CONFIGURATION_DATA_SET_SHEET = "Dataset"

CAVING_PLAN_CONFIGURATION_DATA_SET_AVERAGE_COLUMN = 1
CAVING_PLAN_CONFIGURATION_DATA_SET_SUMMATION_COLUMN = 2


class CavingProductionPlanTarget:
    name: str
    denisty_data_set_name: str
    target_items: list[CavingProductionPlanTargetItem]
    speed_items: list[CavingProductionPlanExtractionSpeedItem]
    average_data_sets: list[str]
    summation_data_sets: list[str]

    def __init__(self, name: str, denisty_data_set_name: str, target_items: list[CavingProductionPlanTargetItem], speed_items: list[CavingProductionPlanExtractionSpeedItem], average_data_sets: list[str], summation_data_sets: list[str]):
        self.name = name
        self.denisty_data_set_name = denisty_data_set_name
        target_items.sort(key=lambda x: x.period_number)
        self.target_items = target_items

        speed_items.sort(key=lambda x: x.minimum_percentage)
        self.speed_items = speed_items

        self.average_data_sets = average_data_sets
        self.summation_data_sets = summation_data_sets

    def export_to_excel(self, filepath: str):
        workbook = Workbook()

        # -- Metadata
        worksheet = workbook.create_sheet(
            CAVING_PLAN_CONFIGURATION_METADATA_SHEET, 0)

        cell = worksheet.cell(
            CAVING_PLAN_CONFIGURATION_NAME_CELL[0], CAVING_PLAN_CONFIGURATION_NAME_CELL[1] - 1)
        cell.value = 'Name'
        cell.font = Font(bold=True)

        cell = worksheet.cell(
            CAVING_PLAN_CONFIGURATION_NAME_CELL[0], CAVING_PLAN_CONFIGURATION_NAME_CELL[1])
        cell.value = self.name

        # Number of periods
        cell = worksheet.cell(
            CAVING_PLAN_CONFIGURATION_PERIOD_CELL[0], CAVING_PLAN_CONFIGURATION_PERIOD_CELL[1] - 1)
        cell.value = 'Number of periods'
        cell.font = Font(bold=True)

        cell = worksheet.cell(
            CAVING_PLAN_CONFIGURATION_PERIOD_CELL[0], CAVING_PLAN_CONFIGURATION_PERIOD_CELL[1])
        cell.value = len(self.target_items)

        # Speed items
        cell = worksheet.cell(
            CAVING_PLAN_CONFIGURATION_NUMBER_OF_SPEEDS_CELL[0], CAVING_PLAN_CONFIGURATION_NUMBER_OF_SPEEDS_CELL[1] - 1)
        cell.value = 'Speed discretization'
        cell.font = Font(bold=True)

        cell = worksheet.cell(
            CAVING_PLAN_CONFIGURATION_NUMBER_OF_SPEEDS_CELL[0], CAVING_PLAN_CONFIGURATION_NUMBER_OF_SPEEDS_CELL[1])
        cell.value = len(self.speed_items)

        # Density
        cell = worksheet.cell(
            CAVING_PLAN_CONFIGURATION_DENSITY_DATA_SET_CELL[0], CAVING_PLAN_CONFIGURATION_DENSITY_DATA_SET_CELL[1] - 1)
        cell.value = 'Density'
        cell.font = Font(bold=True)

        cell = worksheet.cell(
            CAVING_PLAN_CONFIGURATION_DENSITY_DATA_SET_CELL[0], CAVING_PLAN_CONFIGURATION_DENSITY_DATA_SET_CELL[1])
        cell.value = self.denisty_data_set_name

        # Average Data Sets
        cell = worksheet.cell(
            CAVING_PLAN_CONFIGURATION_AVERAGE_DATA_SET_CELL[0], CAVING_PLAN_CONFIGURATION_AVERAGE_DATA_SET_CELL[1] - 1)
        cell.value = 'Number of average items'
        cell.font = Font(bold=True)

        cell = worksheet.cell(
            CAVING_PLAN_CONFIGURATION_AVERAGE_DATA_SET_CELL[0], CAVING_PLAN_CONFIGURATION_AVERAGE_DATA_SET_CELL[1])
        cell.value = len(self.average_data_sets)

        # Summation Data Sets
        cell = worksheet.cell(
            CAVING_PLAN_CONFIGURATION_SUMMATION_DATA_SET_CELL[0], CAVING_PLAN_CONFIGURATION_SUMMATION_DATA_SET_CELL[1] - 1)
        cell.value = 'Number of summation items'
        cell.font = Font(bold=True)

        cell = worksheet.cell(
            CAVING_PLAN_CONFIGURATION_SUMMATION_DATA_SET_CELL[0], CAVING_PLAN_CONFIGURATION_SUMMATION_DATA_SET_CELL[1])
        cell.value = len(self.summation_data_sets)

        # Target
        worksheet = workbook.create_sheet(
            CAVING_PLAN_CONFIGURATION_TARGET_SHEET, 0)

        cell = worksheet.cell(1, CAVING_PLAN_CONFIGURATION_PERIOD_COLUMN)
        cell.font = Font(bold=True)
        cell.value = 'Period'

        cell = worksheet.cell(1, CAVING_PLAN_CONFIGURATION_DURATION_COLUMN)
        cell.font = Font(bold=True)
        cell.value = 'Duration days'

        cell = worksheet.cell(
            1, CAVING_PLAN_CONFIGURATION_INCORPORATION_COLUMN)
        cell.font = Font(bold=True)
        cell.value = 'Incorporation blocks'

        cell = worksheet.cell(1, CAVING_PLAN_CONFIGURATION_TARGET_COLUMN)
        cell.font = Font(bold=True)
        cell.value = 'Target t'

        for i in np.arange(len(self.target_items)):
            target_item = self.target_items[i]

            cell = worksheet.cell(i+2, CAVING_PLAN_CONFIGURATION_PERIOD_COLUMN)
            cell.value = target_item.period_number

            cell = worksheet.cell(i+2, CAVING_PLAN_CONFIGURATION_TARGET_COLUMN)
            cell.value = target_item.target_tonnage

            cell = worksheet.cell(
                i+2, CAVING_PLAN_CONFIGURATION_INCORPORATION_COLUMN)
            cell.value = target_item.incorporation_blocks

            cell = worksheet.cell(i+2,
                                  CAVING_PLAN_CONFIGURATION_DURATION_COLUMN)
            cell.value = target_item.duration_days

        # Speed
        worksheet = workbook.create_sheet(
            CAVING_PLAN_CONFIGURATION_SPEED_SHEET, 0)

        cell = worksheet.cell(
            1, CAVING_PLAN_CONFIGURATION_MINIMUM_PERCENTAGE_COLUMN)
        cell.font = Font(bold=True)
        cell.value = 'Minimum %'

        cell = worksheet.cell(
            1, CAVING_PLAN_CONFIGURATION_MAXIMUM_PERCENTAGE_COLUMN)
        cell.font = Font(bold=True)
        cell.value = 'Maximum %'

        cell = worksheet.cell(
            1, CAVING_PLAN_CONFIGURATION_SPEED_COLUMN)
        cell.font = Font(bold=True)
        cell.value = 'Speed t/d/m²'

        for i in np.arange(len(self.speed_items)):
            speed_item = self.speed_items[i]

            cell = worksheet.cell(
                i+2, CAVING_PLAN_CONFIGURATION_MINIMUM_PERCENTAGE_COLUMN)
            cell.value = speed_item.minimum_percentage

            cell = worksheet.cell(
                i+2, CAVING_PLAN_CONFIGURATION_MAXIMUM_PERCENTAGE_COLUMN)
            cell.value = speed_item.maximum_percentage

            cell = worksheet.cell(
                i+2, CAVING_PLAN_CONFIGURATION_SPEED_COLUMN)
            cell.value = speed_item.extraction_tonnes_per_day_squared_meters

        # Add the density set
        worksheet = workbook.create_sheet(
            CAVING_PLAN_CONFIGURATION_DATA_SET_SHEET, 0)

        cell = worksheet.cell(
            1, CAVING_PLAN_CONFIGURATION_DATA_SET_AVERAGE_COLUMN)
        cell.font = Font(bold=True)
        cell.value = 'Average sets'

        for i in np.arange(len(self.average_data_sets)):
            cell = worksheet.cell(
                2 + i, CAVING_PLAN_CONFIGURATION_DATA_SET_AVERAGE_COLUMN)
            cell.value = self.average_data_sets[i]

        cell = worksheet.cell(
            1, CAVING_PLAN_CONFIGURATION_DATA_SET_SUMMATION_COLUMN)
        cell.font = Font(bold=True)
        cell.value = 'Summation sets'

        for i in np.arange(len(self.summation_data_sets)):
            cell = worksheet.cell(
                2 + i, CAVING_PLAN_CONFIGURATION_DATA_SET_SUMMATION_COLUMN)
            cell.value = self.summation_data_sets[i]

        # Remove the default sheet
        remove_default_worksheet(workbook)
        workbook.save(filepath)
