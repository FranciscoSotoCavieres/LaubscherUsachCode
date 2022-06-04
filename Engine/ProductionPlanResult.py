import numpy as np
from Models.excel_utils import remove_default_worksheet
import Models.utils as utils
from Models import utils
from Engine.ExtractionPeriodBasicScheduleResult import ExtractionPeriodBasicScheduleResult
from Models.BlockModel import BlockModel
from Engine.CavingProductionPlanTarget import CavingProductionPlanTarget
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

PRODUCTION_PLAN_KEYWORD = 'Production plan'


class ProductionPlanResultPeriod:
    period_id: int
    tonnage: float
    active_area_squared_meters: float
    incorporated_area_squared_meters: float
    depleted_area_squared_meters: float
    average: dict[str, float]
    summation: dict[str, float]

    def __init__(self, units: list[ExtractionPeriodBasicScheduleResult], period_id: int, density_set: str,
                 block_model: BlockModel, average_sets: list[str] = None, summation_sets: list[str] = None):

        self.average = dict()
        self.summation = dict()
        self.tonnage = 0
        self.period_id = period_id

        valid_units: list[ExtractionPeriodBasicScheduleResult] = []

        for unit in units:
            if (unit.period_id == period_id):
                self.tonnage = self.tonnage + unit.extracted_tonnage
                valid_units.append(unit)

        density_data = block_model.get_data_set(density_set)

        block_height = block_model.structure.block_size[2]

        if (average_sets != None):
            for average_set in average_sets:
                data_set = block_model.get_data_set(average_set)

                density_values = np.zeros([len(valid_units)])
                average_values = np.zeros([len(valid_units)])

                count = 0

                for unit in valid_units:
                    subscript_i = unit.footprint_subscripts.i
                    subscript_j = unit.footprint_subscripts.j
                    column_data = data_set[subscript_i, subscript_j, :]
                    column_density = density_data[subscript_i, subscript_j, :]
                    initial_fraction: float = unit.from_meters / block_height
                    final_fraction: float = unit.to_meters / block_height
                    average_values[count] = utils.get_average(
                        initial_fraction, final_fraction, column_data, column_density)
                    density_values[count] = utils.get_average(
                        unit.from_meters / block_height, unit.to_meters / block_height, column_density, column_density)

                    count = count + 1
                if (density_values.sum() != 0):
                    average_value = np.average(
                        average_values, weights=density_values)
                    self.average[average_set] = average_value
                else:
                    self.average[average_set] = 0

        if (summation_sets != None):
            for summation_set in summation_sets:
                data_set = block_model.get_data_set(summation_set)

                summation_values = np.zeros([len(valid_units)])

                count = 0
                for unit in valid_units:
                    subscript_i = unit.footprint_subscripts.i
                    subscript_j = unit.footprint_subscripts.j
                    column_data = data_set[subscript_i, subscript_j, :]

                    initial_fraction: float = unit.from_meters / block_height
                    final_fraction: float = unit.to_meters / block_height
                    summation_values[count] = utils.get_summation(
                        initial_fraction, final_fraction, column_data)

                    count = count + 1

                average_value: float = np.sum(summation_values)
                self.summation[summation_set] = average_value

        # Areas
        area_per_unit_square_meters = block_model.structure.get_block_area()

        # Compute the active area
        self.active_area_squared_meters = 0
        for unit in valid_units:
            if unit.extracted_tonnage > 0:
                self.active_area_squared_meters = self.active_area_squared_meters + \
                    area_per_unit_square_meters

        # Compute the incorporated area
        self.incorporated_area_squared_meters = 0
        for unit in valid_units:
            (i, j) = unit.footprint_subscripts.i, unit.footprint_subscripts.j
            subscript_units = (x for x in units if (
                x.footprint_subscripts.i == i and x.footprint_subscripts.j == j))
            # Check if the minimum
            period_ids = [
                subscript_unit.period_id for subscript_unit in subscript_units]
            min_period = min(period_ids)
            if (min_period == period_id):
                self.incorporated_area_squared_meters = self.incorporated_area_squared_meters + \
                    area_per_unit_square_meters

        # Compute the depleted area
        last_month_active_area_squared_meters = 0.0
        for unit in units:
            if (unit.period_id != period_id - 1):
                continue
            if unit.extracted_tonnage > 0:
                last_month_active_area_squared_meters = last_month_active_area_squared_meters + \
                    area_per_unit_square_meters
        self.depleted_area_squared_meters = -(
            self.active_area_squared_meters - self.incorporated_area_squared_meters) + last_month_active_area_squared_meters


class ProductionPlanResult:
    block_model: BlockModel
    units: list[ExtractionPeriodBasicScheduleResult]
    target: CavingProductionPlanTarget

    period_results: dict[int, ProductionPlanResultPeriod]

    average_sets: list[str]
    summation_sets: list[str]

    def __init__(self, units: list[ExtractionPeriodBasicScheduleResult], target: CavingProductionPlanTarget,
                 block_model: BlockModel,
                 average_sets: list[str] = None, summation_sets: list[str] = None):
        self.units = units
        self.block_mode = block_model
        self.average_sets = average_sets
        self.summation_sets = summation_sets
        self.target = target

        # Assign items per period
        self.period_results = dict()
        density_data_set_name = self.target.denisty_data_set_name
        for target_item in self.target.target_items:
            period_id = target_item.period_number
            self.period_results[target_item.period_number] = ProductionPlanResultPeriod(units=self.units,
                                                                                        period_id=period_id,
                                                                                        density_set=density_data_set_name,
                                                                                        average_sets=self.average_sets,
                                                                                        summation_sets=summation_sets,
                                                                                        block_model=block_model)

    def dump_units(self, filepath: str):
        """Dump all units

        Args:
            filepath (str): csv filepath
        """

        lines: list[str] = []

        header = "Period,Subscript I, Subscript J,Extracted Tonnage, From Meters,To Meters,Is Depleted,Tonnage Available, Target Tonnage, Accomplished\n"
        lines.append(header)
        for unit in self.units:
            line = f"{unit.period_id},{unit.footprint_subscripts.i},{unit.footprint_subscripts.j},"
            line = line + \
                f"{unit.extracted_tonnage},{unit.from_meters},{unit.to_meters},{unit.is_depleted},"
            line = line + \
                f"{unit.tonnage_available},{unit.target_tonnage},{unit.was_target_accomplished}\n"
            lines.append(line)

        with open(filepath, 'w+') as write_file:
            write_file.writelines(lines)

    def export_excel(self, filepath: str):
        """Export excel
        Args:
            filepath (str): xlsx filepath
        """
        workbook = Workbook()

        worksheet: Worksheet = workbook.create_sheet(PRODUCTION_PLAN_KEYWORD)

        cell = worksheet.cell(1, 1)

        period_column: int = 1
        tonnage_column: int = 2
        active_area_column: int = 3
        incorporated_area_column: int = 4
        depleted_area_column: int = 5
        data_sets_begin_column: int = 6

        current_row = 1
        worksheet.cell(current_row, period_column).value = 'Period'
        worksheet.cell(current_row, tonnage_column).value = 'Tonnage'
        worksheet.cell(
            current_row, active_area_column).value = 'Active area m²'
        worksheet.cell(
            current_row, incorporated_area_column).value = 'Incorporated area m²'
        worksheet.cell(
            current_row, depleted_area_column).value = 'Depleted area m²'
        current_column = data_sets_begin_column
        for average_set in self.average_sets:
            worksheet.cell(current_row, current_column).value = average_set
            current_column = current_column + 1
        for summation_set in self.summation_sets:
            worksheet.cell(current_row, current_column).value = summation_set
            current_column = current_column + 1

        current_row = current_row + 1
        for period in self.period_results.keys():
            period_result = self.period_results[period]

            worksheet.cell(
                current_row, period_column).value = period_result.period_id
            worksheet.cell(
                current_row, tonnage_column).value = period_result.tonnage
            worksheet.cell(
                current_row, active_area_column).value = period_result.active_area_squared_meters
            worksheet.cell(
                current_row, incorporated_area_column).value = period_result.incorporated_area_squared_meters
            worksheet.cell(
                current_row, depleted_area_column).value = period_result.depleted_area_squared_meters

            current_column = data_sets_begin_column
            for average_set in self.average_sets:
                worksheet.cell(
                    current_row, current_column).value = period_result.average[average_set]
                current_column = current_column + 1

            for summation_set in self.summation_sets:
                worksheet.cell(
                    current_row, current_column).value = period_result.summation[summation_set]
                current_column = current_column + 1

            current_row = current_row + 1
        remove_default_worksheet(workbook)
        workbook.save(filepath)
