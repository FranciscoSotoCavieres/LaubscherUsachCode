from locale import normalize
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl import Workbook, worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side, BORDER_THICK
from openpyxl.styles.fills import PatternFill
from matplotlib.colors import Colormap, to_hex
import math
from Constants.UsachColors import COLOR_2

CELL_SIZE = 3


def get_border(border_type=BORDER_THICK, color: str = COLOR_2):
    side = Side(border_style=border_type, color=color)
    border = Border(left=side, right=side, bottom=side, top=side)
    return border


def get_color(normalized: float, colormap: Colormap) -> PatternFill:

    color_hex: str = to_hex(colormap(int(normalized*255)))[1:]
    color_hex = color_hex.upper()
    return PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')


def export_matrix(array2d: np.ndarray, workbook: Workbook, sheet_name: str, none_value: float = None, cmap='jet'):
    # matplotlib.

    colormap = plt.get_cmap(cmap)

    if (len(array2d[array2d != none_value]) == 0):
        normalized = np.zeros(array2d.shape)
    else:
        if (none_value == None):
            normalized = (array2d - array2d.min()) / \
                (array2d.max()-array2d.min())
        else:
            normalized = (array2d - array2d[array2d != none_value].min())/(
                array2d[array2d != none_value].max()-array2d[array2d != none_value].min())

    worksheet = workbook.create_sheet(sheet_name, 0)

    none_color = openpyxl.styles.PatternFill(fill_type=None)
    # Feed the indices
    m_blocks = array2d.shape[0]
    n_blocks = array2d.shape[1]
    for i in np.arange(m_blocks):
        for j in np.arange(n_blocks):
            cell = worksheet.cell(n_blocks - j, i + 1)
            cell.border = get_border()
            value = array2d[i, j]
            normalized_value = normalized[i, j]

            if (value == None):
                cell.value = value
                cell.fill = none_color

            else:
                if (value == none_value or math.isnan(normalized_value)):
                    cell.value = None
                    cell.fill = none_color
                else:
                    cell.value = value
                    cell.fill = get_color(normalized_value, colormap)

    for j in np.arange(m_blocks):
        worksheet.column_dimensions[get_column_letter(
            j + 1)].width = CELL_SIZE

    for i in np.arange(n_blocks):
        worksheet.row_dimensions[i + 1].ht = CELL_SIZE * 6


def load_matrix(workbook: Workbook, sheet_name: str, rows: int, columns=int):
    sheet: worksheet = workbook[sheet_name]
    array = np.zeros([rows, columns])
    for i in np.arange(rows):
        for j in np.arange(columns):
            value = sheet.cell(columns - j, i + 1).value
            array[i, j] = value
    return array


def remove_default_worksheet(workbook: Workbook):
    """
    It has to have at least 1 other worksheet
    :param workbook: Workbook
    """
    workbook.remove(workbook['Sheet'])



def export_matrix_to_excel(array2d: np.ndarray ,excel_path:str):
    workbook = Workbook()
    export_matrix(array2d,workbook,"matrix")
    workbook.save(excel_path)